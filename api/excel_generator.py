from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


def generate_overtime_report(employee_data: dict, overtimes: list) -> io.BytesIO:
    """
    Генерирует Excel отчёт по переработкам сотрудника

    Args:
        employee_data: словарь с данными сотрудника
        overtimes: список переработок

    Returns:
        BytesIO с Excel файлом
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по переработкам"

    # ========== Стили ==========
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Arial', size=10)

    title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # ========== Заголовок ==========
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"ОТЧЁТ ПО ПЕРЕРАБОТКАМ СОТРУДНИКА"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align

    # ========== Информация о сотруднике ==========
    ws.merge_cells('A3:F3')
    ws['A3'].value = "ИНФОРМАЦИЯ О СОТРУДНИКЕ"
    ws['A3'].font = Font(name='Arial', size=12, bold=True)

    employee_info = [
        ("ФИО:", employee_data.get('full_name', '')),
        ("Должность:", employee_data.get('post', '')),
        ("Отдел:", employee_data.get('office', '')),
        ("Email:", employee_data.get('email', '')),
        ("Телефон:", employee_data.get('phone', '')),
        ("Всего часов переработок:", employee_data.get('total_hours', 0)),
        ("Доступно для использования:", employee_data.get('remaining_hours', 0)),
    ]

    row = 4
    for label, value in employee_info:
        ws.cell(row=row, column=1, value=label).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # ========== Таблица переработок ==========
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="ДЕТАЛИЗАЦИЯ ПЕРЕРАБОТОК").font = Font(name='Arial', size=12, bold=True)

    row += 1
    headers = ["№", "Дата", "Название", "Часы", "Статус", "Накоплено"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ========== Данные ==========
    cumulative_hours = 0

    for i, overtime in enumerate(overtimes, 1):
        row += 1
        cumulative_hours += overtime.get('overtime_hours', 0)

        values = [
            i,
            overtime.get('date_overtime', ''),
            overtime.get('name_overtime', ''),
            overtime.get('overtime_hours', 0),
            overtime.get('status', 'Активен'),
            cumulative_hours
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col in [1, 4, 6]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # ========== Итоговая строка ==========
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    total_cell = ws.cell(row=row, column=1, value="ИТОГО:")
    total_cell.font = Font(name='Arial', size=11, bold=True)
    total_cell.fill = total_fill
    total_cell.alignment = Alignment(horizontal='right', vertical='center')

    total_hours_cell = ws.cell(row=row, column=4, value=employee_data.get('total_hours', 0))
    total_hours_cell.font = Font(name='Arial', size=11, bold=True)
    total_hours_cell.fill = total_fill
    total_hours_cell.alignment = center_align

    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border

    # ========== Дата создания отчёта ==========
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value=f"Отчёт создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(
        name='Arial', size=9, italic=True, color='666666'
    )

    # ========== Настройка ширины колонок ==========
    column_widths = {
        'A': 6,  # №
        'B': 15,  # Дата
        'C': 35,  # Название
        'D': 10,  # Часы
        'E': 15,  # Статус
        'F': 15,  # Накоплено
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ========== Сохраняем в BytesIO ==========
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output